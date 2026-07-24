from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from directory.models import Dealer, Category
import re


class Command(BaseCommand):
    help = 'Import dealers from Excel file with automatic category assignment'

    def add_arguments(self, parser):
        parser.add_argument(
            'file',
            type=str,
            help='Path to Excel file (e.g., RenSetu_Dealers_Formatted.xlsx)',
        )
        parser.add_argument(
            '--skip-existing',
            action='store_true',
            help='Skip dealers that already exist (by name)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview import without saving to database',
        )

    def handle(self, file, *args, **options):
        skip_existing = options.get('skip_existing', False)
        dry_run = options.get('dry_run', False)

        try:
            wb = load_workbook(file)
            ws = wb.active
        except FileNotFoundError:
            raise CommandError(f'File not found: {file}')
        except Exception as e:
            raise CommandError(f'Error reading Excel file: {e}')

        categories_map = {cat.name: cat for cat in Category.objects.all()}

        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        self.stdout.write(
            self.style.HTTP_INFO(f'\n📊 Starting import from: {file}')
        )
        if dry_run:
            self.stdout.write(
                self.style.WARNING('🔍 DRY RUN MODE — no changes will be saved\n')
            )

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name, categories_str, city, area, phone, whatsapp, description, since, verified, active, address, website, rating = row

            if not name or not isinstance(name, str):
                continue

            name = name.strip()

            is_verified = self._parse_bool(verified, default=False)
            is_active = self._parse_bool(active, default=True)

            phone_clean = self._clean_phone(phone)
            whatsapp_clean = self._clean_phone(whatsapp)

            desc_parts = [str(description).strip()] if description else []
            if website and isinstance(website, str):
                website = website.strip()
                if website and website.lower() not in ('none', 'n/a', ''):
                    desc_parts.append(f'Website: {website}')
            final_description = ' '.join(desc_parts) or 'N/A'

            dealer_data = {
                'city': (city or '').strip() if city else '',
                'area': (area or '').strip() if area else '',
                'phone': phone_clean,
                'whatsapp': whatsapp_clean,
                'description': final_description,
                'since': str(since).strip() if since else '',
                'is_verified': is_verified,
                'is_active': is_active,
                'address': (address or '').strip() if address else '',
            }

            dealer_exists = Dealer.objects.filter(name=name).exists()
            if dealer_exists and skip_existing:
                skipped_count += 1
                self.stdout.write(
                    self.style.WARNING(f'⏭️  Row {row_num}: Skipped (exists) — {name}')
                )
                continue

            try:
                if dry_run:
                    self.stdout.write(f'✓ Row {row_num}: Would create — {name}')
                else:
                    dealer, created = Dealer.objects.update_or_create(
                        name=name,
                        defaults=dealer_data,
                    )

                    if categories_str and isinstance(categories_str, str):
                        categories_str = categories_str.strip()
                        if categories_str in categories_map:
                            category = categories_map[categories_str]
                            dealer.categories.add(category)
                            cat_status = f' → {categories_str}'
                        else:
                            cat_status = f' ⚠️  Category "{categories_str}" not found'
                            if categories_str not in [e for e, _, _ in errors]:
                                errors.append((categories_str, 'Category not in database', row_num))
                    else:
                        cat_status = ' (no category)'

                    status = '✨ Created' if created else '🔄 Updated'
                    self.stdout.write(f'{status}: Row {row_num} — {name}{cat_status}')

                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

            except Exception as e:
                error_msg = f'Row {row_num}: {name} — {str(e)}'
                errors.append((name, str(e), row_num))
                self.stdout.write(self.style.ERROR(f'❌ {error_msg}'))

        self.stdout.write('\n' + '=' * 60)
        if dry_run:
            self.stdout.write(self.style.SUCCESS(f'✓ DRY RUN: Would import {created_count + updated_count} dealers'))
        else:
            self.stdout.write(self.style.SUCCESS(
                f'✓ Import complete: {created_count} created, {updated_count} updated, {skipped_count} skipped'
            ))

        if errors:
            self.stdout.write(self.style.WARNING(f'\n⚠️  {len(errors)} issue(s):'))
            for item, error, row in errors:
                self.stdout.write(f'  Row {row}: {item} — {error}')

    def _parse_bool(self, value, default=False):
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ('yes', 'y', 'true', '1', 'active')
        return default

    def _clean_phone(self, phone):
        if not phone:
            return ''
        phone_str = str(phone).strip()
        phone_str = phone_str.replace(' ', '')
        return phone_str or ''